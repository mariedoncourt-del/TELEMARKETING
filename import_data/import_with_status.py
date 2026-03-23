import openpyxl
import json
import re
import os
import sys

# ============================================================
# MAPPING de statuts Excel vers statuts de l'application
# AR = À Rappeler, RDV = Rendez-vous, FIN = Clôturé
# NRP/REPONDEUR/INDISPO/BRS/RAPPEL = prospect déjà contacté mais pas de suite
#   -> on les met en NOUVEAU pour qu'ils restent dans la file
# Pas de statut = NOUVEAU (jamais contacté)
# ============================================================
STATUS_MAP = {
    'AR': 'AR',
    'RDV': 'RDV',
    'FIN': 'FIN',
    'NRP': 'NOUVEAU',       # NRP = juste pas répondu
    'BRS': 'FIN',           # BRS = barré = fini
    'REPONDEUR': 'NOUVEAU', # Répondeur = pas joint
    'RAPPEL': 'AR',         # Rappel = à rappeler
    'INDISPO': 'NOUVEAU',   # Indispo = pas joint
}

def normalize_phone(phone):
    """Normalize French phone numbers"""
    if not phone:
        return None
    phone = str(phone).strip()
    # Handle float format from Excel: 561890870.0 -> 0561890870
    if '.' in phone:
        try:
            phone = str(int(float(phone)))
        except:
            pass
    phone = re.sub(r'[^\d+]', '', phone)
    if phone.startswith('33') and len(phone) >= 11:
        phone = '0' + phone[2:]
    if phone.startswith('335') and len(phone) >= 12:
        phone = '0' + phone[3:]
    if len(phone) == 9 and phone[0] != '0':
        phone = '0' + phone
    if len(phone) != 10:
        return None
    if not phone.startswith(('01','02','03','04','05','06','07','08','09')):
        return None
    return f"{phone[:2]} {phone[2:4]} {phone[4:6]} {phone[6:8]} {phone[8:10]}"

def detect_status_column(headers, rows):
    """Detect the primary status column index"""
    # Priority 1: header explicitly named STATUT
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if 'statut' in h_lower:
            return i
    
    # Priority 2: column with most AR/RDV/FIN values
    best_col = None
    best_count = 0
    for i in range(min(len(headers), 30)):
        count = 0
        for row in rows[:500]:
            if i < len(row) and row[i]:
                val = str(row[i]).strip().upper()
                if val in ('AR', 'RDV', 'FIN', 'NRP', 'BRS'):
                    count += 1
        if count > best_count:
            best_count = count
            best_col = i
    
    if best_count >= 3:
        return best_col
    return None

def detect_comment_column(headers, status_col_idx):
    """Detect the comment/notes column"""
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if any(kw in h_lower for kw in ['commentaire', 'remarque', 'resume', 'notes', 'observation']):
            return i
    # Often the column right after status
    if status_col_idx is not None:
        candidates = [status_col_idx + 1, status_col_idx + 2]
        for c in candidates:
            if c < len(headers):
                h_lower = headers[c].lower().strip()
                if h_lower in ('', 'commentaire', 'remarque', 'notes'):
                    return c
    return None

# ============================================================
# FILE-SPECIFIC EXTRACTION CONFIGS
# Each config tells us how to extract data from each file format
# ============================================================
FILE_CONFIGS = {
    '81-vetements-2025.xlsx': {
        'sheets': ['81-vetements'],
        'mapping': {
            'entreprise': 0,   # raison sociale
            'dirigeant': 1,
            'adresse': 2,
            'ville': 3,
            'code_postal': 5,
            'telephone': 7,
            'status_col': 22,  # col with NRP/REPONDEUR/RDV values
        },
        'opco': 'AGEFICE',
        'source': '81-vetements-2025',
    },
    'Artisans-Eligibles-FAFCEA.xlsx': {
        'sheets': ['Sheet1'],
        'mapping': {
            'siren': 0,
            'entreprise': 1,
            'dirigeant': 2,
            'telephone': 12,
            'code_postal': 8,
            'ville': 10,
            'adresse': 16,    # ville, cp
            'status_col': 13,
            'comment_col': 15,
            'date_appel_col': 14,
        },
        'opco': 'FAFCEA',
        'source': 'Artisans-Eligibles-FAFCEA',
    },
    'Boulangeries.xlsx': {
        'sheets': ['Worksheet'],
        'mapping': {
            'siren': 0,
            'entreprise': 1,
            'telephone': 2,
            'ville': 4,
            'code_ape': 5,
            'dirigeant_prenom': 8,
            'dirigeant_nom': 9,
            'code_postal': 11,
            'adresse': 13,
            'status_col': 22,
            'comment_col': 23,
        },
        'opco': 'AGEFICE',
        'source': 'Boulangeries',
    },
    'ESTHETIQUE-31.xlsx': {
        'sheets': ['Photographes'],  # main sheet with 3776 rows
        'mapping': {
            'entreprise': 1,
            'dirigeant': 2,
            'telephone': 3,
            'adresse': 4,
            'code_postal': 5,
            'ville': 6,
            'code_ape': 11,
            'status_col': 12,
            'date_appel_col': 13,
            'comment_col': 14,
        },
        'opco': 'FAFCEA',
        'source': 'ESTHETIQUE-31',
    },
    'FAFCEA-Artisans-81-82.xlsx': {
        'sheets': ['ESTHETIQUE 31', 'Feuille 3'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'telephone2': 5,
            'status_col': 6,
            'date_appel_col': 8,
            'comment_col': 9,
        },
        'sheet_overrides': {
            'Feuille 3': {
                'status_col': 9,
                'date_appel_col': 10,
                'comment_col': 11,
            }
        },
        'opco': 'FAFCEA',
        'source': 'FAFCEA-Artisans-81-82',
    },
    'IMPRIMERIE-FAFCEA.xlsx': {
        'sheets': ['JARDINERIE 31'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'telephone2': 5,
            'status_col': 13,
            'comment_col': 14,
            'date_appel_col': 16,
        },
        'opco': 'FAFCEA',
        'source': 'IMPRIMERIE-FAFCEA',
    },
    'OPTIQUE-31-2023.xlsx': {
        'sheets': ['OPTIQUE  31- 2018'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'date_appel_col': 5,
            'status_col': 6,
            'comment_col': 7,
        },
        'opco': 'FAFCEA',
        'source': 'OPTIQUE-31-2023',
    },
    'QUINQUAILLERIES-81-2025.xlsx': {
        'sheets': ['81-resto'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'date_appel_col': 5,
            'status_col': 6,
            'comment_col': 7,
        },
        'opco': 'AGEFICE',
        'source': 'QUINQUAILLERIES-81-2025',
    },
    'RESTAURANT-31.xlsx': {
        'sheets': ['81-resto'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'date_appel_col': 5,
            'status_col': 6,
            'comment_col': 7,
        },
        'opco': 'AGEFICE',
        'source': 'RESTAURANT-31',
    },
    'RESTAURANT-81-gsheet.xlsx': {
        'sheets': ['Feuille 1'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'code_postal': 2,
            'ville': 3,
            'telephone': 4,
            'status_col': 5,
            'comment_col': 7,
            'date_appel_col': 8,
        },
        'opco': 'AGEFICE',
        'source': 'RESTAURANT-81-gsheet',
    },
    'SALON-DE-THE-31-2024.xlsx': {
        'sheets': ['SALON DE THE 31'],
        'mapping': {
            'entreprise': 0,
            'adresse': 1,
            'ville': 3,
            'telephone': 4,
            'date_appel_col': 7,
            'status_col': 8,
            'comment_col': 9,
        },
        'opco': 'FAFCEA',
        'source': 'SALON-DE-THE-31-2024',
    },
    'commerces-gros-AKTO.xlsx': {
        'sheets': ['Worksheet'],
        'mapping': {
            'siren': 0,
            'entreprise': 1,
            'dirigeant': 2,
            'telephone': 3,
            'email': 4,
            'ville': 5,
            'code_postal': 10,
            'adresse': 11,
            'status_col': 12,
            'date_appel_col': 13,
            'comment_col': 14,
        },
        'opco': 'AKTO',
        'source': 'commerces-gros-AKTO',
    },
    'fafcea-ALIMENTAIRE.xlsx': {
        'sheets': ['iqualif'],
        'mapping': {
            'code_postal': 1,
            'ville': 2,
            'adresse': 3,
            'entreprise': 4,
            'telephone': 5,
            'status_col': 7,  # col 7 has the actual values
            'date_appel_col': 8,
            'comment_col': 9,
        },
        'opco': 'FAFCEA',
        'source': 'fafcea-ALIMENTAIRE',
    },
    'magasin-chaussures-2025.xlsx': {
        'sheets': ['Worksheet'],
        'mapping': {
            'entreprise': 0,
            'siren': 2,
            'ville': 4,
            'adresse': 5,
            'dirigeant': 6,
            'telephone': 7,
            'email': 8,
            'status_col': 12,
            'comment_col': 13,
        },
        'opco': 'FAFCEA',
        'source': 'magasin-chaussures-2025',
    },
    'commerces-gros-AKTO-6000E.xlsx': {
        'sheets': ['Worksheet'],
        'mapping': {
            'entreprise': 0,
            'siren': 2,
            'ville': 7,
            'code_postal': 8,
            'dirigeant': 15,
            'telephone': 16,
            'email': 17,
            'status_col': 11,  # col 11 has "commentaires"/repondeur/nrp
            'comment_col': 11,
        },
        'opco': 'AKTO',
        'source': 'commerces-gros-AKTO-6000E',
    },
}

def get_cell_value(row, idx):
    """Safely get a cell value from a row"""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ''
    val = str(row[idx]).strip()
    if val.lower() in ('none', 'nan', 'null'):
        return ''
    return val

def extract_status(raw_value):
    """Extract status from raw cell value"""
    if not raw_value:
        return 'NOUVEAU'
    val = raw_value.strip().upper()
    # Direct match
    if val in STATUS_MAP:
        return STATUS_MAP[val]
    # Check if starts with a known status
    for k, v in STATUS_MAP.items():
        if val.startswith(k):
            return v
    # Special cases
    if 'DÉCROCHÉ' in val or 'DECROCHE' in val:
        return 'NOUVEAU'  # Just picked up, no real status
    if 'PAS INTERE' in val or 'PAS INTÉRÉ' in val or 'RACCROCH' in val:
        return 'FIN'
    if 'RAPPEL' in val or 'RAPL' in val or 'RAPP' in val:
        return 'AR'
    return 'NOUVEAU'

def extract_nrp_count(status_raw, comment):
    """Try to extract NRP count from status and comments"""
    count = 0
    combined = f"{status_raw} {comment}".upper()
    
    # "NRP 3 FOIS" or "3 FOIS NE REPONDS"
    match = re.search(r'(\d+)\s*(?:FOIS|X)', combined)
    if match:
        count = max(count, int(match.group(1)))
    
    if 'NRP' in combined.upper():
        count = max(count, 1)
    if 'REPONDEUR' in combined.upper():
        count = max(count, 1)
    
    return count

def process_file(filename, config):
    """Process a single Excel file"""
    filepath = os.path.join('.', filename)
    if not os.path.exists(filepath):
        print(f"  SKIP: {filename} not found")
        return []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ERROR: {filename}: {e}")
        return []
    
    prospects = []
    mapping = config['mapping']
    target_sheets = config.get('sheets', wb.sheetnames)
    
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: Sheet '{sheet_name}' not found in {filename}")
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        
        # Apply sheet-specific overrides if any
        effective_mapping = dict(mapping)
        sheet_overrides = config.get('sheet_overrides', {})
        if sheet_name in sheet_overrides:
            effective_mapping.update(sheet_overrides[sheet_name])
        
        # Determine if first row is headers
        has_headers = any(
            h and any(kw in str(h).lower() for kw in ['nom', 'raison', 'tel', 'adresse', 'ville', 'statut', 'siren', 'entreprise', 'code'])
            for h in rows[0] if h
        )
        
        start_row = 1 if has_headers else 0
        
        for row in rows[start_row:]:
            # Extract phone
            phone_col = effective_mapping.get('telephone')
            if phone_col is None:
                # Try to find phone in various cols
                for ci in range(min(len(row), 20)):
                    val = get_cell_value(row, ci)
                    if normalize_phone(val):
                        phone_col = ci
                        break
            
            phone = normalize_phone(get_cell_value(row, phone_col)) if phone_col is not None else None
            # Try telephone2 as fallback
            if not phone and effective_mapping.get('telephone2') is not None:
                phone = normalize_phone(get_cell_value(row, effective_mapping['telephone2']))
            if not phone:
                continue
            
            # Extract enterprise name
            entreprise = get_cell_value(row, effective_mapping.get('entreprise'))
            if not entreprise:
                entreprise = 'INCONNUE'
            # Clean enterprise name
            entreprise = entreprise[:200]
            
            # Extract status
            status_raw = get_cell_value(row, effective_mapping.get('status_col'))
            statut = extract_status(status_raw)
            
            # Extract comment
            comment = get_cell_value(row, effective_mapping.get('comment_col', None))
            
            # Combine status_raw into comment if it contains useful info
            if status_raw and status_raw.upper() not in ('AR', 'RDV', 'FIN', 'NRP', 'BRS', 'REPONDEUR', 'RAPPEL', 'INDISPO', 'NOUVEAU', ''):
                # Status contains extra info like "repondeur ar"
                if comment:
                    comment = f"{status_raw} | {comment}"
                else:
                    comment = status_raw
            
            # NRP count
            nrp_count = extract_nrp_count(status_raw, comment)
            
            # Build prospect
            prospect = {
                'entreprise': entreprise,
                'dirigeant': get_cell_value(row, effective_mapping.get('dirigeant')) or 
                             (f"{get_cell_value(row, effective_mapping.get('dirigeant_prenom'))} {get_cell_value(row, effective_mapping.get('dirigeant_nom'))}".strip() if effective_mapping.get('dirigeant_prenom') else ''),
                'telephone': phone,
                'email': get_cell_value(row, effective_mapping.get('email')),
                'adresse': get_cell_value(row, effective_mapping.get('adresse')),
                'code_postal': get_cell_value(row, effective_mapping.get('code_postal')),
                'ville': get_cell_value(row, effective_mapping.get('ville')) or 'INCONNUE',
                'code_ape': get_cell_value(row, effective_mapping.get('code_ape')),
                'siren': get_cell_value(row, effective_mapping.get('siren')),
                'opco': config.get('opco', 'NON_DETERMINE'),
                'source': config.get('source', filename),
                'statut': statut,
                'notes': comment[:500] if comment else '',
                'nrp_count': nrp_count,
            }
            
            prospects.append(prospect)
    
    wb.close()
    return prospects


# ============================================================
# MAIN
# ============================================================
all_prospects = []
stats = {'total': 0, 'by_file': {}, 'by_status': {}, 'by_opco': {}}

for filename, config in FILE_CONFIGS.items():
    prospects = process_file(filename, config)
    if prospects:
        stats['by_file'][filename] = {
            'total': len(prospects),
            'statuses': {}
        }
        for p in prospects:
            s = p['statut']
            stats['by_file'][filename]['statuses'][s] = stats['by_file'][filename]['statuses'].get(s, 0) + 1
            stats['by_status'][s] = stats['by_status'].get(s, 0) + 1
            stats['by_opco'][p['opco']] = stats['by_opco'].get(p['opco'], 0) + 1
        
        all_prospects.extend(prospects)
        print(f"  {filename}: {len(prospects)} prospects ({', '.join(f'{k}:{v}' for k,v in stats['by_file'][filename]['statuses'].items())})")

# Deduplicate by phone
seen_phones = set()
unique_prospects = []
for p in all_prospects:
    if p['telephone'] not in seen_phones:
        seen_phones.add(p['telephone'])
        unique_prospects.append(p)

stats['total'] = len(unique_prospects)
stats['duplicates_removed'] = len(all_prospects) - len(unique_prospects)

print(f"\n{'='*60}")
print(f"TOTAL UNIQUE: {stats['total']} prospects ({stats['duplicates_removed']} doublons supprimés)")
print(f"PAR STATUT: {json.dumps(stats['by_status'], indent=2)}")
print(f"PAR OPCO: {json.dumps(stats['by_opco'], indent=2)}")

# Save to JSON
with open('prospects_import_v2.json', 'w', encoding='utf-8') as f:
    json.dump(unique_prospects, f, ensure_ascii=False, indent=None)

print(f"\nSauvegardé dans prospects_import_v2.json")
